"""Analytics: Excel экспорт"""

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ...schemas.analytics import AnalyticsResponse


def _sanitize_sheet_name(name: str) -> str:
    """Санитизирует имя листа для Excel (<=31 символ, без запрещённых символов)"""
    # Удаляем запрещённые символы
    sanitized = re.sub(r'[\\/:?*\[\]]', '_', name)
    # Обрезаем до 31 символа
    if len(sanitized) > 31:
        sanitized = sanitized[:31]
    return sanitized


def _add_kpi_sheet(wb: Workbook, kpis: list) -> None:
    """Добавляет лист KPI"""
    ws = wb.create_sheet('KPI')

    # Заголовки
    headers = ['Key', 'Value', 'Unit', 'Delta', 'Delta Dir', 'Caption']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

    # Данные
    for row, kpi in enumerate(kpis, 2):
        ws.cell(row=row, column=1, value=kpi.key)
        ws.cell(row=row, column=2, value=kpi.value)
        ws.cell(row=row, column=3, value=kpi.unit)
        ws.cell(row=row, column=4, value=kpi.delta)
        ws.cell(row=row, column=5, value=kpi.delta_dir)
        ws.cell(row=row, column=6, value=kpi.caption)

    # Автоширина колонок
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _add_chart_sheet(wb: Workbook, chart_idx: int, chart: Any) -> None:
    """Добавляет лист для графика"""
    title_safe = _sanitize_sheet_name(chart.title)
    sheet_name = f'Chart-{chart_idx}-{title_safe}'[:31]
    ws = wb.create_sheet(sheet_name)

    # Заголовок
    ws.cell(row=1, column=1, value=f'Chart: {chart.title}')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f'Type: {chart.type}')

    # Разворачиваем data в плоскую таблицу в зависимости от типа
    if chart.type == 'line':
        # points: [{date, value}]
        points = chart.data.get('points', [])
        ws.cell(row=4, column=1, value='Date')
        ws.cell(row=4, column=2, value='Value')
        for i, point in enumerate(points, 5):
            ws.cell(row=i, column=1, value=point.get('date'))
            ws.cell(row=i, column=2, value=point.get('value'))

    elif chart.type == 'bar' or chart.type == 'hbar':
        # items: [{label, value}] или [{recruiter, value}]
        items = chart.data.get('items', [])
        headers = list(items[0].keys()) if items else ['label', 'value']
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header.capitalize())
        for i, item in enumerate(items, 5):
            for col, key in enumerate(headers, 1):
                ws.cell(row=i, column=col, value=item.get(key))

    elif chart.type == 'heatmap':
        # cells: [{x, y, value}], x_labels, y_labels
        cells = chart.data.get('cells', [])
        x_labels = chart.data.get('x_labels', [])
        y_labels = chart.data.get('y_labels', [])

        ws.cell(row=4, column=1, value='Y Label')
        ws.cell(row=4, column=2, value='X Label')
        ws.cell(row=4, column=3, value='Value')

        for i, cell in enumerate(cells, 5):
            x_idx = cell.get('x', 0)
            y_idx = cell.get('y', 0)
            x_label = x_labels[x_idx] if x_idx < len(x_labels) else str(x_idx)
            y_label = y_labels[y_idx] if y_idx < len(y_labels) else str(y_idx)

            ws.cell(row=i, column=1, value=y_label)
            ws.cell(row=i, column=2, value=x_label)
            ws.cell(row=i, column=3, value=cell.get('value'))

    elif chart.type == 'cohort':
        # cohorts: [{month, sizes: [{day, retained_pct}]}]
        cohorts = chart.data.get('cohorts', [])
        ws.cell(row=4, column=1, value='Month')
        ws.cell(row=4, column=2, value='Day')
        ws.cell(row=4, column=3, value='Retained Pct')

        row_idx = 5
        for cohort in cohorts:
            month = cohort.get('month')
            for size in cohort.get('sizes', []):
                ws.cell(row=row_idx, column=1, value=month)
                ws.cell(row=row_idx, column=2, value=size.get('day'))
                ws.cell(row=row_idx, column=3, value=size.get('retained_pct'))
                row_idx += 1

    elif chart.type == 'scatter':
        # points: [{label, x, y}]
        points = chart.data.get('points', [])
        ws.cell(row=4, column=1, value='Label')
        ws.cell(row=4, column=2, value='X')
        ws.cell(row=4, column=3, value='Y')
        for i, point in enumerate(points, 5):
            ws.cell(row=i, column=1, value=point.get('label'))
            ws.cell(row=i, column=2, value=point.get('x'))
            ws.cell(row=i, column=3, value=point.get('y'))

    elif chart.type == 'pie':
        # our: [{reason, value, pct}], candidate: [{reason, value, pct}]
        ws.cell(row=4, column=1, value='Side')
        ws.cell(row=4, column=2, value='Reason')
        ws.cell(row=4, column=3, value='Value')
        ws.cell(row=4, column=4, value='Pct')

        row_idx = 5
        for side, data in chart.data.items():
            if isinstance(data, list):
                for item in data:
                    ws.cell(row=row_idx, column=1, value=side)
                    ws.cell(row=row_idx, column=2, value=item.get('reason'))
                    ws.cell(row=row_idx, column=3, value=item.get('value'))
                    ws.cell(row=row_idx, column=4, value=item.get('pct'))
                    row_idx += 1

    elif chart.type == 'funnel':
        # stages: [{stage_key, label, color, count, conversion_from_prev_pct}]
        stages = chart.data.get('stages', [])
        terminals = chart.data.get('terminals', {})

        ws.cell(row=4, column=1, value='Stage Key')
        ws.cell(row=4, column=2, value='Label')
        ws.cell(row=4, column=3, value='Count')
        ws.cell(row=4, column=4, value='Conversion Pct')

        row_idx = 5
        for stage in stages:
            ws.cell(row=row_idx, column=1, value=stage.get('stage_key'))
            ws.cell(row=row_idx, column=2, value=stage.get('label'))
            ws.cell(row=row_idx, column=3, value=stage.get('count'))
            ws.cell(row=row_idx, column=4, value=stage.get('conversion_from_prev_pct'))
            row_idx += 1

        # Terminals
        if terminals:
            row_idx += 1
            ws.cell(row=row_idx, column=1, value='Terminals:')
            row_idx += 1
            for term_key, term_data in terminals.items():
                ws.cell(row=row_idx, column=1, value=term_key)
                ws.cell(row=row_idx, column=2, value=term_data.get('n'))
                ws.cell(row=row_idx, column=3, value=f"{term_data.get('pct')}%")
                row_idx += 1

    elif chart.type == 'stacked':
        # Различные структуры в зависимости от отчёта
        if 'stages' in chart.data:
            # Overview: stages: [{stage_key, label, color, count}]
            stages = chart.data.get('stages', [])
            ws.cell(row=4, column=1, value='Stage Key')
            ws.cell(row=4, column=2, value='Label')
            ws.cell(row=4, column=3, value='Count')
            for i, stage in enumerate(stages, 5):
                ws.cell(row=i, column=1, value=stage.get('stage_key'))
                ws.cell(row=i, column=2, value=stage.get('label'))
                ws.cell(row=i, column=3, value=stage.get('count'))

        elif 'sources' in chart.data:
            # Sources: sources: [{source, stages: [{stage_key, label, color, count}]}]
            sources = chart.data.get('sources', [])
            ws.cell(row=4, column=1, value='Source')
            ws.cell(row=4, column=2, value='Stage Key')
            ws.cell(row=4, column=3, value='Count')
            row_idx = 5
            for source_data in sources:
                source = source_data.get('source')
                for stage in source_data.get('stages', []):
                    ws.cell(row=row_idx, column=1, value=source)
                    ws.cell(row=row_idx, column=2, value=stage.get('stage_key'))
                    ws.cell(row=row_idx, column=3, value=stage.get('count'))
                    row_idx += 1

    elif chart.type == 'boxplot':
        # stages: [{stage_key, label, median, q1, q3, min, max, outliers}]
        stages = chart.data.get('stages', [])
        ws.cell(row=4, column=1, value='Stage Key')
        ws.cell(row=4, column=2, value='Label')
        ws.cell(row=4, column=3, value='Median')
        ws.cell(row=4, column=4, value='Q1')
        ws.cell(row=4, column=5, value='Q3')
        ws.cell(row=4, column=6, value='Min')
        ws.cell(row=4, column=7, value='Max')
        ws.cell(row=4, column=8, value='Outliers')

        for i, stage in enumerate(stages, 5):
            ws.cell(row=i, column=1, value=stage.get('stage_key'))
            ws.cell(row=i, column=2, value=stage.get('label'))
            ws.cell(row=i, column=3, value=stage.get('median'))
            ws.cell(row=i, column=4, value=stage.get('q1'))
            ws.cell(row=i, column=5, value=stage.get('q3'))
            ws.cell(row=i, column=6, value=stage.get('min'))
            ws.cell(row=i, column=7, value=stage.get('max'))
            outliers = stage.get('outliers', [])
            ws.cell(row=i, column=8, value=str(outliers) if outliers else '')

    elif chart.type == 'survival':
        # points: [{day, retained_pct}]
        points = chart.data.get('points', [])
        ws.cell(row=4, column=1, value='Day')
        ws.cell(row=4, column=2, value='Retained Pct')
        for i, point in enumerate(points, 5):
            ws.cell(row=i, column=1, value=point.get('day'))
            ws.cell(row=i, column=2, value=point.get('retained_pct'))

    elif chart.type == 'radar':
        # axes: [...], series: [{name, values: [...]}]
        axes = chart.data.get('axes', [])
        series = chart.data.get('series', [])

        # Заголовки: Name + оси
        ws.cell(row=4, column=1, value='Name')
        for col, axis in enumerate(axes, 2):
            ws.cell(row=4, column=col, value=axis)

        # Данные
        for i, serie in enumerate(series, 5):
            ws.cell(row=i, column=1, value=serie.get('name'))
            values = serie.get('values', [])
            for col, value in enumerate(values, 2):
                ws.cell(row=i, column=col, value=value)

    # Автоширина
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _add_table_sheet(wb: Workbook, table_idx: int, table: Any) -> None:
    """Добавляет лист для таблицы"""
    title_safe = _sanitize_sheet_name(table.title)
    sheet_name = f'Table-{table_idx}-{title_safe}'[:31]
    ws = wb.create_sheet(sheet_name)

    # Заголовок
    ws.cell(row=1, column=1, value=f'Table: {table.title}')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Заголовки колонок
    columns = table.columns
    for col, column in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col, value=column.label)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

    # Данные
    rows = table.rows
    for row_idx, row_data in enumerate(rows, 4):
        for col, column in enumerate(columns, 1):
            value = row_data.get(column.key)
            ws.cell(row=row_idx, column=col, value=value)

    # Автоширина
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def build_xlsx(report_name: str, response: AnalyticsResponse) -> bytes:
    """Строит Excel файл из AnalyticsResponse"""
    wb = Workbook()

    # Удаляем стандартный лист
    default_sheet = wb.active
    wb.remove(default_sheet)

    # KPI лист
    if response.kpis:
        _add_kpi_sheet(wb, response.kpis)

    # Charts
    for idx, chart in enumerate(response.charts, 1):
        _add_chart_sheet(wb, idx, chart)

    # Tables
    for idx, table in enumerate(response.tables, 1):
        _add_table_sheet(wb, idx, table)

    # Если нет листов, создаём пустой
    if len(wb.worksheets) == 0:
        wb.create_sheet('Empty')

    # Сохраняем в память
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()