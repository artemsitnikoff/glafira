"""Сервис импорта кандидатов из Excel файлов"""

import asyncio
import json
import logging
import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional, Dict, List, Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select, or_, func, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from ..core.errors import ValidationError
from ..database import AsyncSessionLocal
from ..models import (
    Candidate,
    CandidateImportJob,
    CandidateExperience,
    CandidateSkill,
    CandidateEducation,
    Comment,
    Consent,
)
from ..schemas.candidate import CandidateSource
from ..services.audit import audit
from ..services.integrations.potok.client import get_all_applicants, preview_applicants, iter_applicants
from ..services.integrations.potok.mapper import map_potok_applicant
from ..services.integrations.talantix.client import TalantixClient, TalantixError
from ..services.integrations.talantix import mapper as talantix_mapper
from .base_search import reindex_all_embeddings
from .candidate_dedup import (
    _clean_phone,
    _normalize_contact,
    _phone_query_variants,
    _get_existing_candidates,
    find_duplicate_candidates,
)
from .phone import normalize_phone

logger = logging.getLogger(__name__)

# Удерживаем ссылки на фоновые задачи импорта — иначе event loop держит лишь слабую
# ссылку и GC может убить задачу до завершения (job навсегда застрянет в running).
_active_tasks: dict = {}

# Потолок размера загружаемого файла (защита воркера от OOM на гигантском .xlsx).
MAX_IMPORT_FILE_BYTES = 15 * 1024 * 1024

# Потолок строк против zip-bomb/OOM: сжатый .xlsx из 20 001 строки → отказ с ошибкой.
MAX_IMPORT_ROWS = 20000


def _fit(value, maxlen: int):
    """Обрезает строковое значение под лимит колонки БД (иначе StringDataRightTruncation
    на commit уронит весь батч). None — как есть."""
    if value is None:
        return None
    s = str(value)
    return s[:maxlen] if len(s) > maxlen else s

# Синонимы для авто-распознавания колонок (fuzzy matching, lowercase)
FIELD_SYNONYMS = {
    "name": ["фио кандидата", "фио", "имя", "кандидат", "full name", "полное имя"],
    "phone": ["моб. телефон", "мобильный телефон", "телефон", "номер телефона", "phone", "мобильный"],
    "email": ["e-mail", "эл. почта", "почта", "email", "электронная почта"],
    "city": ["город", "город проживания", "city", "местоположение"],
    "age": ["возраст", "age", "лет"],
    "salary": ["зарплатные ожидания", "зп", "ожидания по зарплате", "salary", "зарплата"],
    "source": ["источник", "источник отклика", "по способу добавления", "source"],
    "position": ["желаемая должность", "должность", "позиция", "position", "роль"],
    "company": ["компания", "место работы", "company", "организация"],
    "experience": ["опыт", "стаж", "experience", "опыт работы"],
    "comment": ["комментарий", "комментарий рекрутёра", "note", "заметка"],
    "resume_url": ["резюме", "ссылка на резюме", "cv link", "резюме ссылка", "cv url"]
}


def _parse_excel_sync(content: bytes) -> tuple[list[str], dict[str, list[str]], int]:
    """Синхронный парсинг Excel файла с возвратом колонок, примеров и количества строк"""

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)

    # Первая строка - заголовки
    headers_row = next(rows_iter, None)
    if not headers_row:
        raise ValidationError("Файл пустой или не содержит заголовков")

    # Очищаем заголовки от None и приводим к строкам
    columns = [str(cell).strip() if cell is not None else f"Колонка {i+1}"
               for i, cell in enumerate(headers_row)]

    # Собираем примеры данных (до 3 на колонку)
    samples = {col: [] for col in columns}
    data_rows = []

    for row in rows_iter:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # Пропускаем пустые строки

        data_rows.append(row)

        if len(data_rows) > MAX_IMPORT_ROWS:
            raise ValidationError(f"Файл превышает лимит {MAX_IMPORT_ROWS} строк — разбейте на части")

        # Собираем примеры для каждой колонки
        for i, cell in enumerate(row):
            if i < len(columns) and cell is not None:
                cell_str = str(cell).strip()
                if cell_str and len(samples[columns[i]]) < 3:
                    samples[columns[i]].append(cell_str)

    workbook.close()
    return columns, samples, len(data_rows)


async def parse_excel_file(content: bytes, filename: str) -> dict:
    """Парсинг Excel файла и возврат метаданных"""
    file_ext = Path(filename).suffix.lower()

    if file_ext == '.xls':
        raise ValidationError("Старый формат .xls не поддерживается. Пересохраните файл в формате .xlsx")

    if file_ext != '.xlsx':
        raise ValidationError("Поддерживаются только файлы формата .xlsx")

    try:
        columns, samples, row_count = await asyncio.to_thread(_parse_excel_sync, content)
    except ValidationError:
        raise
    except Exception as e:
        logger.error(f"Ошибка парсинга Excel файла {filename}: {e}")
        raise ValidationError("Не удалось прочитать файл. Поддерживаются только .xlsx файлы")

    # Автоматическое распознавание колонок
    auto_mapping = _auto_map_columns(columns)

    return {
        "columns": columns,
        "samples": samples,
        "row_count": row_count,
        "auto_mapping": auto_mapping
    }


def _auto_map_columns(columns: list[str]) -> dict[str, str]:
    """Автоматическое распознавание колонок файла"""
    mapping = {}
    used_fields = set()

    for col in columns:
        col_lower = col.lower().strip()

        # Ищем лучшее совпадение
        best_field = None
        for field_key, synonyms in FIELD_SYNONYMS.items():
            if field_key in used_fields:
                continue

            for synonym in synonyms:
                if synonym in col_lower:
                    best_field = field_key
                    break

            if best_field:
                break

        if best_field:
            mapping[col] = best_field
            used_fields.add(best_field)
        else:
            mapping[col] = "skip"

    return mapping


def _clean_name(value: str) -> tuple[str, str, str]:
    """Очистка и разбор ФИО. Возвращает (last_name, first_name, middle_name)"""
    if not value:
        return "", "", ""

    # Убираем коды анонимизации типа "062-" (только в начале токена — чтобы не клеить
    # слова вроде "Иван123-Петров").
    cleaned = re.sub(r'\b\d{3}-', '', str(value)).strip()

    # Разбиваем по пробелам
    parts = [p.strip() for p in cleaned.split() if p.strip()]

    if len(parts) == 0:
        return "", "", ""
    elif len(parts) == 1:
        return "", parts[0], ""  # first_name
    elif len(parts) == 2:
        return parts[0], parts[1], ""  # last_name, first_name
    else:
        # 3+ токена: фамилия, имя, остальное как отчество
        return parts[0], parts[1], " ".join(parts[2:])




def _parse_company_position(value: str, field_type: str) -> tuple[str, str]:
    """Парсинг значения 'Компания: Должность' - возвращает (company, position)"""
    if not value or ':' not in value:
        if field_type == 'company':
            return str(value or ""), ""
        else:  # position
            return "", str(value or "")

    parts = value.split(':', 1)
    company = parts[0].strip()
    position = parts[1].strip()

    return company, position


def _clean_salary(value: str) -> Optional[int]:
    """Извлечение зарплаты как числа"""
    if not value:
        return None

    # Убираем пробелы и нецифровые символы, оставляем только цифры
    digits = re.sub(r'[^\d]', '', str(value))

    if not digits:
        return None

    try:
        salary = int(digits)
        # Санити-проверка
        if salary < 0 or salary > 10_000_000:
            return None
        return salary
    except ValueError:
        return None


def _clean_source(value: str) -> str:
    """Очистка и маппинг источника"""
    if not value:
        return "other"

    value_lower = str(value).lower()

    if any(x in value_lower for x in ['headhunter', 'hh']):
        return "hh"
    elif any(x in value_lower for x in ['вручную', 'manual']):
        return "manual"
    elif any(x in value_lower for x in ['telegram', 'телеграм']):
        return "telegram"
    elif any(x in value_lower for x in ['avito', 'авито']):
        return "avito"
    elif 'superjob' in value_lower:
        return "superjob"
    else:
        return "other"




def _classify_rows(rows: list[dict], existing_candidates: list[Candidate],
                  mapping: dict[str, str]) -> list[dict]:
    """Классификация строк на new/duplicate/error"""
    # Создаем словари существующих контактов
    existing_phones = {_normalize_contact(c.phone or ""): c for c in existing_candidates if c.phone}
    existing_emails = {_normalize_contact(c.email or ""): c for c in existing_candidates if c.email}

    # Для отслеживания внутрифайловых дублей
    seen_phones = {}
    seen_emails = {}

    classified_rows = []

    for i, row in enumerate(rows):
        # Базовая информация
        preview_row = {
            "index": i + 1,
            "name": "",
            "phone": "",
            "email": "",
            "city": "",
            "source": "other",
            "status": "new",
            "error": None,
            "detail": {}
        }

        # Заполняем детали из маппинга
        name_value = ""
        phone_value = ""
        email_value = ""

        for col, field in mapping.items():
            if field == "skip":
                continue

            value = row.get(col, "")
            if not value:
                continue

            if field == "name":
                name_value = str(value).strip()
                last_name, first_name, middle_name = _clean_name(name_value)
                preview_row["name"] = f"{first_name} {last_name}".strip()
                preview_row["detail"]["full_name"] = f"{last_name} {first_name} {middle_name}".strip()
                preview_row["detail"]["last_name"] = last_name
                preview_row["detail"]["first_name"] = first_name
                preview_row["detail"]["middle_name"] = middle_name

            elif field == "phone":
                phone_value = str(value).strip()
                preview_row["phone"] = _clean_phone(phone_value)
                preview_row["detail"]["phone"] = preview_row["phone"]

            elif field == "email":
                email_value = str(value).strip()
                preview_row["email"] = email_value
                preview_row["detail"]["email"] = email_value

            elif field == "city":
                preview_row["city"] = str(value).strip()
                preview_row["detail"]["city"] = str(value).strip()

            elif field == "source":
                preview_row["source"] = _clean_source(str(value))
                preview_row["detail"]["source"] = preview_row["source"]

            elif field == "position":
                company, position = _parse_company_position(str(value), "position")
                if position:
                    preview_row["detail"]["position"] = position
                if company:
                    preview_row["detail"]["company"] = company

            elif field == "company":
                company, position = _parse_company_position(str(value), "company")
                if company:
                    preview_row["detail"]["company"] = company
                if position:
                    preview_row["detail"]["position"] = position

            elif field == "experience":
                preview_row["detail"]["experience"] = str(value).strip()

            elif field == "age":
                try:
                    age = int(str(value))
                    preview_row["detail"]["age"] = age
                except (ValueError, TypeError):
                    pass

            elif field == "salary":
                salary = _clean_salary(str(value))
                if salary:
                    preview_row["detail"]["salary"] = salary

            elif field == "comment":
                preview_row["detail"]["comment"] = str(value).strip()

            elif field == "resume_url":
                url = str(value).strip()
                if len(url) <= 500:
                    preview_row["detail"]["resume_url"] = url

        # Проверяем ошибки
        if not name_value.strip():
            preview_row["status"] = "error"
            preview_row["error"] = "нет имени"
        elif not phone_value.strip() and not email_value.strip():
            preview_row["status"] = "error"
            preview_row["error"] = "нет контакта"
        else:
            # Проверяем дедупликацию
            norm_phone = _normalize_contact(phone_value)
            norm_email = _normalize_contact(email_value)

            # Базовый дубль (есть в БД) — запоминаем кандидата для режима «Обновить».
            matched = None
            if norm_phone and norm_phone in existing_phones:
                matched = existing_phones[norm_phone]
            elif norm_email and norm_email in existing_emails:
                matched = existing_emails[norm_email]
            # Внутрифайловый дубль (повтор более ранней строки этого же файла).
            within_file = bool(
                (norm_phone and norm_phone in seen_phones)
                or (norm_email and norm_email in seen_emails)
            )

            if matched is not None:
                preview_row["status"] = "duplicate"
                preview_row["_match_id"] = str(matched.id)
            elif within_file:
                preview_row["status"] = "duplicate"
            else:
                # Запоминаем контакты для следующих строк
                if norm_phone:
                    seen_phones[norm_phone] = i
                if norm_email:
                    seen_emails[norm_email] = i

        classified_rows.append(preview_row)

    return classified_rows


async def preview_import(session: AsyncSession, company_id: UUID, content: bytes,
                        mapping: dict[str, str], dedup_mode: str) -> dict:
    """Превью импорта с классификацией строк"""
    # Парсим файл заново
    columns, samples, row_count = await asyncio.to_thread(_parse_excel_sync, content)

    # Читаем все строки данных

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active

    rows_iter = worksheet.iter_rows(values_only=True)
    next(rows_iter)  # Пропускаем заголовки

    rows = []
    for row in rows_iter:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        # Преобразуем row в словарь по колонкам
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(columns):
                row_dict[columns[i]] = cell
        rows.append(row_dict)

        if len(rows) > MAX_IMPORT_ROWS:
            raise ValidationError(f"Файл превышает лимит {MAX_IMPORT_ROWS} строк — разбейте на части")

    workbook.close()

    # Собираем все контакты для проверки дедупликации
    all_phones = []
    all_emails = []

    for row in rows:
        for col, field in mapping.items():
            value = row.get(col, "")
            if not value:
                continue

            if field == "phone":
                norm_phone = _normalize_contact(_clean_phone(str(value)))
                if norm_phone:
                    all_phones.append(norm_phone)
            elif field == "email":
                norm_email = _normalize_contact(str(value))
                if norm_email:
                    all_emails.append(norm_email)

    # Получаем существующих кандидатов
    existing_candidates = await _get_existing_candidates(session, company_id, all_phones, all_emails)

    # Классифицируем строки
    classified_rows = _classify_rows(rows, existing_candidates, mapping)

    # Подсчитываем статистику
    total = len(classified_rows)
    new = len([r for r in classified_rows if r["status"] == "new"])
    duplicates = len([r for r in classified_rows if r["status"] == "duplicate"])
    errors = len([r for r in classified_rows if r["status"] == "error"])

    # Возвращаем первые 50 строк (без внутренних служебных ключей вроде _match_id)
    shown = min(50, total)
    remaining = max(0, total - shown)
    clean_rows = [
        {k: v for k, v in r.items() if not k.startswith("_")}
        for r in classified_rows[:shown]
    ]

    return {
        "summary": {
            "total": total,
            "new": new,
            "duplicates": duplicates,
            "errors": errors
        },
        "rows": clean_rows,
        "shown": shown,
        "remaining": remaining
    }


async def create_import_job(session: AsyncSession, company_id: UUID, total_rows: int) -> CandidateImportJob:
    """Создание задачи импорта"""
    job = CandidateImportJob(
        company_id=company_id,
        total=total_rows,
        status="running"
    )
    session.add(job)
    await session.flush()
    return job


async def get_import_job(session: AsyncSession, job_id: UUID, company_id: UUID) -> Optional[CandidateImportJob]:
    """Получение задачи импорта"""
    result = await session.execute(
        select(CandidateImportJob).where(
            CandidateImportJob.id == job_id,
            CandidateImportJob.company_id == company_id
        )
    )
    return result.scalar_one_or_none()


def _utc_naive_now() -> datetime:
    """Текущий UTC без tzinfo для finished_at"""
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _apply_update(candidate: Candidate, detail: dict) -> None:
    """Режим «Обновить»: переписывает поля существующего кандидата значениями из файла
    (только непустыми — не затираем имеющееся пустым). Имя/этап не трогаем."""
    if detail.get("phone"):
        candidate.phone = normalize_phone(detail["phone"]) or _fit(detail["phone"], 20)
    if detail.get("email"):
        candidate.email = _fit(detail["email"], 255)
    if detail.get("city"):
        candidate.city = _fit(detail["city"], 120)
    if detail.get("salary"):
        salary_value = detail["salary"]
        candidate.salary_expectation = salary_value
        candidate.salary_from = salary_value
        candidate.salary_to = salary_value
    if detail.get("position"):
        candidate.last_position = _fit(detail["position"], 255)
    if detail.get("company"):
        candidate.last_company = _fit(detail["company"], 255)
    if detail.get("experience"):
        candidate.last_period = _fit(detail["experience"], 120)
    if detail.get("resume_url"):
        candidate.source_url = _fit(detail["resume_url"], 500)


async def _update_job_progress(job_id: UUID, **updates):
    """Обновление прогресса задачи короткой сессией"""

    try:
        async with AsyncSessionLocal() as session:
            job = await session.get(CandidateImportJob, job_id)
            if not job:
                return

            for key, value in updates.items():
                setattr(job, key, value)

            await asyncio.wait_for(session.commit(), timeout=10)
    except Exception as e:
        logger.warning(f"Ошибка обновления прогресса импорта {job_id}: {e}")


async def _finalize_job(job_id: UUID, status: str, error: str = None):
    """Финализация задачи импорта"""

    try:
        async with AsyncSessionLocal() as session:
            job = await session.get(CandidateImportJob, job_id)
            if not job:
                return

            job.status = status
            job.finished_at = _utc_naive_now()
            if error:
                job.error = error

            await asyncio.wait_for(session.commit(), timeout=10)
    except Exception as e:
        logger.error(f"Ошибка финализации импорта {job_id}: {e}")


async def _run_import(job_id: UUID, company_id: UUID, user_id: UUID,
                     file_bytes: bytes, mapping: dict[str, str], dedup_mode: str):
    """Фоновое выполнение импорта"""

    try:
        # Парсим файл
        columns, samples, row_count = await asyncio.to_thread(_parse_excel_sync, file_bytes)

        # Читаем все строки
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        worksheet = workbook.active

        rows_iter = worksheet.iter_rows(values_only=True)
        next(rows_iter)  # Пропускаем заголовки

        all_rows = []
        for row in rows_iter:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(columns):
                    row_dict[columns[i]] = cell
            all_rows.append(row_dict)

            if len(all_rows) > MAX_IMPORT_ROWS:
                workbook.close()
                await _finalize_job(
                    job_id, "error",
                    f"Файл превышает лимит {MAX_IMPORT_ROWS} строк — разбейте на части"
                )
                return

        workbook.close()

        # Обновляем total
        await _update_job_progress(job_id, total=len(all_rows))

        # Получаем существующих кандидатов для дедупликации
        all_phones = []
        all_emails = []

        for row in all_rows:
            for col, field in mapping.items():
                value = row.get(col, "")
                if not value:
                    continue

                if field == "phone":
                    norm_phone = _normalize_contact(_clean_phone(str(value)))
                    if norm_phone:
                        all_phones.append(norm_phone)
                elif field == "email":
                    norm_email = _normalize_contact(str(value))
                    if norm_email:
                        all_emails.append(norm_email)

        async with AsyncSessionLocal() as session:
            existing_candidates = await _get_existing_candidates(session, company_id, all_phones, all_emails)

        # Классифицируем строки
        classified_rows = _classify_rows(all_rows, existing_candidates, mapping)

        # Обрабатываем батчами
        batch_size = 500
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors_count = 0

        for i in range(0, len(classified_rows), batch_size):
            batch = classified_rows[i:i + batch_size]

            async with AsyncSessionLocal() as session:
                for row_data in batch:
                    try:
                        if row_data["status"] == "error":
                            errors_count += 1
                            continue

                        # Каждую строку обрабатываем в savepoint: сбой одной записи
                        # (нарушение CHECK и т.п.) откатывает ТОЛЬКО её, не «отравляет»
                        # сессию и не валит весь батч (до 500 кандидатов). Счётчики
                        # инкрементим строго ПОСЛЕ успешного flush.
                        async with session.begin_nested():
                            if row_data["status"] == "duplicate":
                                match_id = row_data.get("_match_id")
                                if dedup_mode == "update" and match_id:
                                    # Реальное обновление существующего кандидата полями из файла.
                                    existing = await session.get(Candidate, UUID(match_id))
                                    if existing is not None and existing.company_id == company_id:
                                        _apply_update(existing, row_data["detail"])
                                        await session.flush()
                                        updated_count += 1
                                    else:
                                        skipped_count += 1
                                else:
                                    # skip-режим ИЛИ внутрифайловый дубль (без базового совпадения)
                                    skipped_count += 1
                            elif row_data["status"] == "new":
                                detail = row_data["detail"]
                                candidate = Candidate(
                                    company_id=company_id,  # ВСЕГДА из контекста, не из файла
                                    last_name=_fit(detail.get("last_name") or "", 120),
                                    first_name=_fit(detail.get("first_name") or "Неизвестно", 120),
                                    middle_name=_fit(detail.get("middle_name") or None, 120),
                                    phone=normalize_phone(detail.get("phone")) or _fit(detail.get("phone"), 20),
                                    email=_fit(detail.get("email"), 255),
                                    city=_fit(detail.get("city"), 120),
                                    salary_expectation=detail.get("salary"),
                                    salary_from=detail.get("salary"),
                                    salary_to=detail.get("salary"),
                                    last_position=_fit(detail.get("position"), 255),
                                    last_company=_fit(detail.get("company"), 255),
                                    last_period=_fit(detail.get("experience"), 120),
                                    source=detail.get("source") or "import",
                                    external_source="import",
                                    source_url=_fit(detail.get("resume_url"), 500),
                                    extra={"imported": True}
                                )
                                session.add(candidate)
                                await session.flush()
                                created_count += 1

                    except Exception as e:
                        logger.error(f"Ошибка обработки строки импорта Excel: {e}")
                        errors_count += 1

                await session.commit()

            # Обновляем прогресс
            processed = min(i + batch_size, len(classified_rows))
            await _update_job_progress(
                job_id,
                processed=processed,
                created=created_count,
                updated=updated_count,
                skipped=skipped_count,
                errors=errors_count
            )

        # Создаем audit запись
        async with AsyncSessionLocal() as session:
            await audit(
                session,
                action="candidates_import",
                entity_type="candidate_import_job",
                entity_id=job_id,
                after={
                    "created": created_count,
                    "updated": updated_count,
                    "skipped": skipped_count,
                    "errors": errors_count
                },
                actor_type="human",
                actor_user_id=user_id,
                company_id=company_id
            )
            await session.commit()

        # Финализируем
        await _finalize_job(job_id, "done")

        # Запускаем переиндексацию эмбеддингов для всей компании
        try:
            await reindex_all_embeddings(company_id)
            logger.info(f"Запущена переиндексация эмбеддингов после импорта для компании {company_id}")
        except Exception as e:
            logger.error(f"Ошибка запуска переиндексации после импорта: {e}")

    except Exception as e:
        logger.error(f"Ошибка выполнения импорта {job_id}: {e}")
        await _finalize_job(job_id, "error", str(e)[:500])


async def execute_import(session: AsyncSession, company_id: UUID, user_id: UUID,
                        content: bytes, mapping: dict[str, str], dedup_mode: str) -> UUID:
    """Запуск импорта в фоновой задаче"""
    # Считаем общее количество строк
    columns, samples, row_count = await asyncio.to_thread(_parse_excel_sync, content)

    # Создаем задачу
    job = await create_import_job(session, company_id, row_count)
    await session.commit()

    # Запускаем в фоне. Удерживаем ссылку в _active_tasks — иначе GC может убить задачу
    # до завершения (job застрянет в running навсегда). callback снимает ссылку по завершении.
    task = asyncio.create_task(_run_import(job.id, company_id, user_id, content, mapping, dedup_mode))
    _active_tasks[job.id] = task

    def _cleanup_task(_t, _jid=job.id):
        _active_tasks.pop(_jid, None)

    task.add_done_callback(_cleanup_task)

    return job.id


# === POTOK IMPORT FUNCTIONS ===

async def preview_potok_import(session: AsyncSession, company_id: UUID, token: str, dedup_mode: str) -> dict:
    """
    Превью импорта кандидатов из Potok.io с дедупликацией

    Использует LIGHT preview_applicants() — быстрый запрос без полной загрузки

    Args:
        session: DB сессия
        company_id: ID компании
        token: API токен Potok
        dedup_mode: режим дедупликации ('skip' или 'update')

    Returns:
        Статистика превью аналогично Excel-импорту
    """
    logger.info("Начинаем превью импорта Potok")

    try:
        # Быстрый превью без полной загрузки всех ~15,700
        estimated_total, preview_sample = await preview_applicants(token, sample=50)

        logger.info(f"Превью Potok: получено {len(preview_sample)} образцов, оценка общего количества: {estimated_total}")

    except Exception as e:
        logger.error(f"Ошибка превью Potok: {e}")
        raise

    # Маппим кандидатов для превью
    mapped_candidates = []
    for raw in preview_sample:
        try:
            mapped = map_potok_applicant(raw)
            mapped_candidates.append(mapped)
        except Exception as e:
            logger.error(f"Ошибка маппинга кандидата Potok {raw.get('id', '?')}: {e}")
            continue

    # Собираем контакты для дедупликации
    all_phones = []
    all_emails = []
    for candidate in mapped_candidates:
        if candidate.get("phone"):
            norm_phone = _normalize_contact(candidate["phone"])
            if norm_phone:
                all_phones.append(norm_phone)
        if candidate.get("email"):
            norm_email = _normalize_contact(candidate["email"])
            if norm_email:
                all_emails.append(norm_email)

    # Получаем существующих кандидатов
    existing_candidates = await _get_existing_candidates(session, company_id, all_phones, all_emails)

    # Классифицируем кандидатов для превью
    classified_rows = _classify_potok_rows(mapped_candidates, existing_candidates)

    # Подсчет статистики для превью
    preview_total = len(classified_rows)
    preview_new = len([r for r in classified_rows if r["status"] == "new"])
    preview_duplicates = len([r for r in classified_rows if r["status"] == "duplicate"])
    preview_errors = len([r for r in classified_rows if r["status"] == "error"])

    # ДИАГНОСТИКА: откуда дубли (база vs внутри-выгрузки) и сколько кандидатов базы совпало
    _base_dups = sum(1 for r in classified_rows if r.get("_match_id"))
    logger.info(
        f"[potok-preview] sample={len(preview_sample)} existing_matched={len(existing_candidates)} "
        f"new={preview_new} dup={preview_duplicates} (base={_base_dups}, within={preview_duplicates - _base_dups}) "
        f"err={preview_errors}"
    )

    # Экстраполируем статистику на весь объем данных
    if len(preview_sample) > 0 and estimated_total > len(preview_sample):
        scale_factor = estimated_total / len(preview_sample)
        total = int(preview_total * scale_factor)
        new = int(preview_new * scale_factor)
        duplicates = int(preview_duplicates * scale_factor)
        errors = int(preview_errors * scale_factor)
    else:
        total = preview_total
        new = preview_new
        duplicates = preview_duplicates
        errors = preview_errors

    # Возвращаем preview sample
    shown = len(classified_rows)
    remaining = max(0, total - shown)
    clean_rows = [
        {k: v for k, v in r.items() if not k.startswith("_")}
        for r in classified_rows
    ]

    return {
        "summary": {
            "total": total,
            "new": new,
            "duplicates": duplicates,
            "errors": errors
        },
        "rows": clean_rows,
        "shown": shown,
        "remaining": remaining
    }


def _classify_potok_rows(candidates: list[dict], existing_candidates: list[Candidate]) -> list[dict]:
    """Классификация кандидатов Potok на new/duplicate/error (аналог _classify_rows)"""
    # Создаем словари существующих контактов
    existing_phones = {_normalize_contact(c.phone or ""): c for c in existing_candidates if c.phone}
    existing_emails = {_normalize_contact(c.email or ""): c for c in existing_candidates if c.email}

    # Для отслеживания внутренних дублей
    seen_phones = {}
    seen_emails = {}

    classified_rows = []

    for i, candidate in enumerate(candidates):
        # Базовая информация для превью
        first_name = candidate.get("first_name", "")
        last_name = candidate.get("last_name", "")
        # маппер отдаёт phone/email как None (не "") → coerce, иначе .strip() ниже падает
        phone = candidate.get("phone") or ""
        email = candidate.get("email") or ""

        preview_row = {
            "index": i + 1,
            "name": f"{first_name} {last_name}".strip() or "Неизвестно",
            "phone": phone,
            "email": email,
            "city": candidate.get("city", ""),
            "source": "potok",
            "status": "new",
            "error": None,
            "detail": candidate  # полные данные для импорта
        }

        # Проверяем ошибки
        if not first_name.strip() and not last_name.strip():
            preview_row["status"] = "error"
            preview_row["error"] = "нет имени"
        elif not phone.strip() and not email.strip():
            preview_row["status"] = "error"
            preview_row["error"] = "нет контакта"
        else:
            # Проверяем дедупликацию
            norm_phone = _normalize_contact(phone)
            norm_email = _normalize_contact(email)

            # Базовый дубль (есть в БД)
            matched = None
            if norm_phone and norm_phone in existing_phones:
                matched = existing_phones[norm_phone]
            elif norm_email and norm_email in existing_emails:
                matched = existing_emails[norm_email]

            # Внутренний дубль (повтор в этом же запросе)
            within_request = bool(
                (norm_phone and norm_phone in seen_phones)
                or (norm_email and norm_email in seen_emails)
            )

            if matched is not None:
                preview_row["status"] = "duplicate"
                preview_row["_match_id"] = str(matched.id)
            elif within_request:
                preview_row["status"] = "duplicate"
            else:
                # Запоминаем контакты для следующих строк
                if norm_phone:
                    seen_phones[norm_phone] = i
                if norm_email:
                    seen_emails[norm_email] = i

        classified_rows.append(preview_row)

    return classified_rows


def _apply_potok_update(candidate: Candidate, detail: dict) -> None:
    """
    Режим «Обновить» для кандидата Potok: обновляем скалярные поля.

    Для MVP обновляем только основные поля, child-таблицы (опыт/навыки/образование)
    пересоздаем только для НОВЫХ кандидатов чтобы не плодить дубли.
    """
    if detail.get("phone"):
        candidate.phone = normalize_phone(detail["phone"]) or _fit(detail["phone"], 20)
    if detail.get("email"):
        candidate.email = _fit(detail["email"], 255)
    if detail.get("city"):
        candidate.city = _fit(detail["city"], 120)
    if detail.get("birth_date"):
        candidate.birth_date = detail["birth_date"]
    if detail.get("gender"):
        candidate.gender = detail["gender"]
    if detail.get("salary_expectation"):
        salary_value = detail["salary_expectation"]
        candidate.salary_expectation = salary_value
        candidate.salary_from = salary_value
        candidate.salary_to = salary_value
    if detail.get("last_position"):
        candidate.last_position = _fit(detail["last_position"], 255)
    if detail.get("resume_text"):
        candidate.resume_text = detail["resume_text"]
    if detail.get("resume_summary"):
        candidate.resume_summary = detail["resume_summary"]
    if detail.get("source_url"):
        candidate.source_url = _fit(detail["source_url"], 500)
    if detail.get("external_id"):
        candidate.external_id = _fit(detail["external_id"], 120)


async def _create_potok_child_records(session: AsyncSession, company_id: UUID, candidate_id: UUID, detail: dict):
    """Создание связанных записей (опыт, навыки, образование) для кандидата Potok"""

    # Опыт работы
    experience_list = detail.get("experience") or []
    for exp in experience_list:
        if exp.get("position"):
            session.add(CandidateExperience(
                company_id=company_id,
                candidate_id=candidate_id,
                position=_fit(exp["position"], 255),
                company=_fit(exp.get("company"), 255) if exp.get("company") else None,
                period=_fit(exp.get("period"), 120) if exp.get("period") else None,
                description=exp.get("description"),
                order_index=exp.get("order_index", 0)
            ))

    # Навыки
    skills_list = detail.get("skills") or []
    for skill in skills_list:
        if skill.get("skill"):
            session.add(CandidateSkill(
                company_id=company_id,
                candidate_id=candidate_id,
                skill=_fit(skill["skill"], 120),
                order_index=skill.get("order_index", 0)
            ))

    # Образование
    education_list = detail.get("education") or []
    for edu in education_list:
        if edu.get("institution"):
            session.add(CandidateEducation(
                company_id=company_id,
                candidate_id=candidate_id,
                institution=_fit(edu["institution"], 255),
                specialty=_fit(edu.get("specialty"), 255) if edu.get("specialty") else None,
                years=_fit(edu.get("years"), 40) if edu.get("years") else None,
                order_index=edu.get("order_index", 0)
            ))


async def _run_potok_import(job_id: UUID, company_id: UUID, user_id: UUID, token: str, dedup_mode: str):
    """Фоновое выполнение СТРИМИНГОВОГО импорта из Potok.io"""

    try:
        logger.info("Начинаем стриминговый импорт из Potok")

        # Загружаем базовые наборы для дедупликации ОДИН раз
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(Candidate.phone, Candidate.email)
                .where(Candidate.company_id == company_id)
                .where(Candidate.deleted_at.is_(None))
            )
            existing_rows = result.fetchall()

        # Предпроцессинг в наборы для быстрого поиска
        base_phones = set()
        base_emails = set()
        for phone, email in existing_rows:
            if phone:
                norm_phone = _normalize_contact(phone)
                if norm_phone:
                    base_phones.add(norm_phone)
            if email:
                norm_email = _normalize_contact(email)
                if norm_email:
                    base_emails.add(norm_email)

        # Наборы для отслеживания дублей в рамках текущего импорта
        seen_phones = set()
        seen_emails = set()

        # Счётчики
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors_count = 0
        processed_total = 0

        # Progress callback с троттлингом
        def update_progress(done: int, total: int):
            if done == 1 or done % 20 == 0 or done == total:
                asyncio.create_task(_update_job_progress(job_id, processed=done, total=total))

        # STREAMING: обрабатываем каждый батч по мере поступления
        async for batch in iter_applicants(token, on_progress=update_progress):
            batch_mapped = []

            # Маппинг батча
            for raw in batch:
                try:
                    mapped = map_potok_applicant(raw)
                    batch_mapped.append(mapped)
                except Exception as e:
                    logger.error(f"Ошибка маппинга кандидата Potok {raw.get('id', '?')}: {e}")
                    errors_count += 1

            # Классифицируем батч
            for candidate in batch_mapped:
                try:
                    # Проверка на обязательные поля
                    if not candidate.get("first_name") and not candidate.get("last_name"):
                        errors_count += 1
                        continue

                    phone = candidate.get("phone")
                    email = candidate.get("email")

                    if not phone and not email:
                        errors_count += 1
                        continue

                    # Проверка на дублирование
                    norm_phone = _normalize_contact(phone) if phone else None
                    norm_email = _normalize_contact(email) if email else None

                    is_duplicate = False
                    match_id = None

                    # Проверяем дублирование в базе И внутри импорта
                    if norm_phone and (norm_phone in base_phones or norm_phone in seen_phones):
                        is_duplicate = True
                    elif norm_email and (norm_email in base_emails or norm_email in seen_emails):
                        is_duplicate = True

                    # Обрабатываем в зависимости от статуса
                    async with AsyncSessionLocal() as session:
                        async with session.begin_nested():
                            if is_duplicate:
                                if dedup_mode == "update" and match_id:
                                    # TODO: Реализовать поиск match_id для обновления
                                    # Пока пропускаем обновления в стриминговом режиме
                                    skipped_count += 1
                                else:
                                    skipped_count += 1
                            else:
                                # Новый кандидат
                                detail = candidate
                                candidate_obj = Candidate(
                                    company_id=company_id,
                                    first_name=_fit(detail.get("first_name") or "Неизвестно", 120),
                                    last_name=_fit(detail.get("last_name") or "", 120),
                                    middle_name=_fit(detail.get("middle_name"), 120) if detail.get("middle_name") else None,
                                    phone=(normalize_phone(detail.get("phone")) or _fit(detail.get("phone"), 20)) if detail.get("phone") else None,
                                    email=_fit(detail.get("email"), 255) if detail.get("email") else None,
                                    city=_fit(detail.get("city"), 120) if detail.get("city") else None,
                                    birth_date=detail.get("birth_date"),
                                    gender=detail.get("gender"),
                                    salary_expectation=detail.get("salary_expectation"),
                                    salary_from=detail.get("salary_expectation"),
                                    salary_to=detail.get("salary_expectation"),
                                    last_position=_fit(detail.get("last_position"), 255) if detail.get("last_position") else None,
                                    resume_text=detail.get("resume_text"),
                                    resume_summary=detail.get("resume_summary"),
                                    source="potok",
                                    external_source="potok",
                                    external_id=_fit(detail.get("external_id"), 120) if detail.get("external_id") else None,
                                    source_url=_fit(detail.get("source_url"), 500) if detail.get("source_url") else None,
                                    extra={"imported": True, "languages": detail.get("languages") or []}
                                )
                                session.add(candidate_obj)
                                await session.flush()  # получаем ID

                                # Создаем связанные записи и фиксируем их в savepoint
                                await _create_potok_child_records(session, company_id, candidate_obj.id, detail)
                                await session.flush()

                                # Добавляем в seen наборы для внутренней дедупликации
                                if norm_phone:
                                    seen_phones.add(norm_phone)
                                if norm_email:
                                    seen_emails.add(norm_email)

                                created_count += 1

                        await session.commit()

                    processed_total += 1

                except Exception as e:
                    logger.error(f"Ошибка обработки кандидата Potok: {e}")
                    errors_count += 1

            # Обновляем прогресс после каждого батча
            await _update_job_progress(
                job_id,
                processed=processed_total,
                created=created_count,
                updated=updated_count,
                skipped=skipped_count,
                errors=errors_count
            )

        # Создаем audit запись
        async with AsyncSessionLocal() as session:
            await audit(
                session,
                action="candidates_import_potok",
                entity_type="candidate_import_job",
                entity_id=job_id,
                after={
                    "source": "potok",
                    "created": created_count,
                    "updated": updated_count,
                    "skipped": skipped_count,
                    "errors": errors_count
                },
                actor_type="human",
                actor_user_id=user_id,
                company_id=company_id
            )
            await session.commit()

        # Финализируем
        await _finalize_job(job_id, "done")

        # Запускаем переиндексацию эмбеддингов для всей компании
        try:
            await reindex_all_embeddings(company_id)
            logger.info(f"Запущена переиндексация эмбеддингов после Potok импорта для компании {company_id}")
        except Exception as e:
            logger.error(f"Ошибка запуска переиндексации после Potok импорта: {e}")

    except Exception as e:
        logger.error(f"Ошибка выполнения импорта Potok {job_id}: {e}")
        await _finalize_job(job_id, "error", str(e)[:500])


async def execute_potok_import(session: AsyncSession, company_id: UUID, user_id: UUID,
                              token: str, dedup_mode: str) -> UUID:
    """Запуск импорта из Potok в фоновой задаче"""

    # Создаем задачу с предварительным total (будет обновлен в _run_potok_import)
    job = await create_import_job(session, company_id, 0)  # total обновится позже
    await session.commit()

    # Запускаем в фоне с защитой от GC
    task = asyncio.create_task(_run_potok_import(job.id, company_id, user_id, token, dedup_mode))
    _active_tasks[job.id] = task

    def _cleanup_task(_t, _jid=job.id):
        _active_tasks.pop(_jid, None)

    task.add_done_callback(_cleanup_task)

    return job.id


# === TALANTIX IMPORT FUNCTIONS ===
# Зеркало Potok-импорта, но источник — GraphQL API Talantix. ГЛАВНАЯ ЦЕННОСТЬ —
# импорт истории комментариев кандидата (source='talantix') с автором и датой оригинала.

# Кандидаты в Talantix (PersonItem.id) считаем перечислением — total нет.
TALANTIX_PREVIEW_SAMPLE = 50            # сколько персон детально тянуть для оценки new/dup
TALANTIX_PREVIEW_MAX_PAGES = 200        # кап страниц id-перечисления в превью (200*50=10k)
TALANTIX_IMPORT_MAX_PAGES = 4000        # кап страниц id-перечисления в импорте (4000*50=200k)


def _classify_talantix_rows(candidates: list[dict], existing_candidates: list[Candidate]) -> list[dict]:
    """Классификация замапленных персон Talantix на new/duplicate/error (аналог Potok)."""
    existing_phones = {_normalize_contact(c.phone or ""): c for c in existing_candidates if c.phone}
    existing_emails = {_normalize_contact(c.email or ""): c for c in existing_candidates if c.email}

    seen_phones: dict = {}
    seen_emails: dict = {}
    classified_rows = []

    for i, candidate in enumerate(candidates):
        first_name = candidate.get("first_name") or ""
        last_name = candidate.get("last_name") or ""
        phone = candidate.get("phone") or ""
        email = candidate.get("email") or ""

        preview_row = {
            "index": i + 1,
            "name": f"{first_name} {last_name}".strip() or "Неизвестно",
            "phone": phone,
            "email": email,
            "city": candidate.get("city") or "",
            "source": "talantix",
            "status": "new",
            "error": None,
            "detail": candidate,
        }

        if not first_name.strip() and not last_name.strip():
            preview_row["status"] = "error"
            preview_row["error"] = "нет имени"
        elif not phone.strip() and not email.strip():
            preview_row["status"] = "error"
            preview_row["error"] = "нет контакта"
        else:
            norm_phone = _normalize_contact(phone)
            norm_email = _normalize_contact(email)

            matched = None
            if norm_phone and norm_phone in existing_phones:
                matched = existing_phones[norm_phone]
            elif norm_email and norm_email in existing_emails:
                matched = existing_emails[norm_email]

            within_request = bool(
                (norm_phone and norm_phone in seen_phones)
                or (norm_email and norm_email in seen_emails)
            )

            if matched is not None:
                preview_row["status"] = "duplicate"
                preview_row["_match_id"] = str(matched.id)
            elif within_request:
                preview_row["status"] = "duplicate"
            else:
                if norm_phone:
                    seen_phones[norm_phone] = i
                if norm_email:
                    seen_emails[norm_email] = i

        classified_rows.append(preview_row)

    return classified_rows


async def preview_talantix_import(session: AsyncSession, company_id: UUID, dedup_mode: str) -> dict:
    """Превью импорта из Talantix: total (перечислением, капнуто) + сэмпл new/dup/errors.

    Токен уже сохранён (connect). Здесь только читаем: перечисляем id персон (id-only,
    легко), тянем полные карточки первых N для дедуп-классификации, экстраполируем."""
    logger.info("Начинаем превью импорта Talantix company=%s", company_id)

    async with TalantixClient(company_id) as client:
        person_ids = await client.collect_person_ids(max_pages=TALANTIX_PREVIEW_MAX_PAGES)
        total = len(person_ids)

        sample_ids = person_ids[:TALANTIX_PREVIEW_SAMPLE]
        detail_results = await asyncio.gather(
            *[client.get_person_full(pid) for pid in sample_ids],
            return_exceptions=True,
        )

    mapped_candidates = []
    for res in detail_results:
        if isinstance(res, Exception) or res is None:
            continue
        try:
            mapped_candidates.append(talantix_mapper.map_person(res))
        except Exception as e:  # noqa: BLE001
            logger.error("Ошибка маппинга персоны Talantix для превью: %s", e)

    # Дедуп сэмпла
    all_phones = []
    all_emails = []
    for c in mapped_candidates:
        if c.get("phone"):
            np = _normalize_contact(c["phone"])
            if np:
                all_phones.append(np)
        if c.get("email"):
            ne = _normalize_contact(c["email"])
            if ne:
                all_emails.append(ne)

    existing_candidates = await _get_existing_candidates(session, company_id, all_phones, all_emails)
    classified_rows = _classify_talantix_rows(mapped_candidates, existing_candidates)

    sample_total = len(classified_rows)
    sample_new = len([r for r in classified_rows if r["status"] == "new"])
    sample_dup = len([r for r in classified_rows if r["status"] == "duplicate"])
    sample_err = len([r for r in classified_rows if r["status"] == "error"])

    # Экстраполяция на весь объём (как в Potok)
    if sample_total > 0 and total > sample_total:
        scale = total / sample_total
        summary = {
            "total": total,
            "new": int(sample_new * scale),
            "duplicates": int(sample_dup * scale),
            "errors": int(sample_err * scale),
        }
    else:
        summary = {
            "total": total,
            "new": sample_new,
            "duplicates": sample_dup,
            "errors": sample_err,
        }

    # Превью упёрлось в кап перечисления person_ids → реальная база БОЛЬШЕ,
    # чем показано (фронт рисует «N+», чтобы не дискредитировать реальный объём).
    summary["total_capped"] = total >= TALANTIX_PREVIEW_MAX_PAGES * 50

    shown = len(classified_rows)
    clean_rows = [
        {k: v for k, v in r.items() if not k.startswith("_")}
        for r in classified_rows
    ]
    return {
        "summary": summary,
        "rows": clean_rows,
        "shown": shown,
        "remaining": max(0, total - shown),
    }


async def _generate_consent_number(session: AsyncSession, company_id: UUID) -> str:
    """Сгенерировать app-native номер согласия (PD-{seq}/{yy}) с advisory-lock (как в consent.py).

    Формат совместим с генератором в services/consent.py (та же последовательность),
    чтобы ручное создание согласий не ломалось на CAST разбора номера."""
    await session.execute(
        text("SELECT pg_advisory_xact_lock(hashtext(:k))"),
        {"k": f"consent_number:{company_id}"},
    )
    yy = str(datetime.now(timezone.utc).year)[-2:]
    seq = (
        await session.execute(
            text(
                "SELECT COALESCE(MAX(CAST(SPLIT_PART(SPLIT_PART(number, '-', 2), '/', 1) AS INTEGER)), 0) + 1 "
                "FROM consents WHERE company_id = :cid AND number LIKE :pat"
            ),
            {"cid": company_id, "pat": f"PD-%/{yy}"},
        )
    ).scalar_one()
    return f"PD-{seq:03d}/{yy}"


def _apply_talantix_update(candidate: Candidate, mapped: dict) -> None:
    """Режим «Обновить»: обогащаем скалярные поля существующего кандидата (имя не трогаем)."""
    if mapped.get("phone"):
        candidate.phone = normalize_phone(mapped["phone"]) or _fit(mapped["phone"], 20)
    if mapped.get("email"):
        candidate.email = _fit(mapped["email"], 255)
    if mapped.get("city"):
        candidate.city = _fit(mapped["city"], 120)
    if mapped.get("birth_date"):
        candidate.birth_date = mapped["birth_date"]
    if mapped.get("gender"):
        candidate.gender = mapped["gender"]
    if mapped.get("last_position"):
        candidate.last_position = _fit(mapped["last_position"], 255)
    if mapped.get("resume_text"):
        candidate.resume_text = mapped["resume_text"]


async def _resolve_existing_talantix_candidate(
    session: AsyncSession, company_id: UUID, tpid: str | None, phone: str | None, email: str | None
) -> Candidate | None:
    """Найти существующего кандидата: сначала по talantix_person_id, затем по контактам."""
    if tpid:
        res = await session.execute(
            select(Candidate).where(
                Candidate.company_id == company_id,
                Candidate.deleted_at.is_(None),
                Candidate.talantix_person_id == tpid,
            )
        )
        c = res.scalars().first()
        if c is not None:
            return c
    dupes = await find_duplicate_candidates(session, company_id, phone, email)
    return dupes[0] if dupes else None


async def _import_talantix_comments(
    session: AsyncSession, company_id: UUID, candidate_id: UUID, events: list[dict]
) -> int:
    """Импорт истории-комментариев (CommentAdded/HhCommentAdded) → Comment(source='talantix').

    Дедуп по (company_id, source='talantix', external_id). Автор и дата ОРИГИНАЛА
    сохраняются (author_name_ext / created_at), НЕ приписываются импортёру.
    Каждая вставка в savepoint — гонка крон↔on-demand не валит батч."""
    imported = 0
    now = datetime.now(timezone.utc)
    for node in events:
        mapped = talantix_mapper.map_comment_event(node)
        if not mapped:
            continue
        ext_id = _fit(mapped["external_id"], 120)
        if not ext_id:
            continue

        existing = (
            await session.execute(
                select(Comment.id).where(
                    Comment.company_id == company_id,
                    Comment.candidate_id == candidate_id,  # dedup СТРОГО в пределах кандидата
                    Comment.source == "talantix",
                    Comment.external_id == ext_id,
                )
            )
        ).scalar_one_or_none()
        if existing:
            continue

        comment = Comment(
            company_id=company_id,
            candidate_id=candidate_id,
            application_id=None,
            author_user_id=None,  # у комментария из Talantix нет нашего автора
            source="talantix",
            external_id=ext_id,
            author_name_ext=mapped.get("author_name"),
            body=mapped["body"],
            mentions=[],
            created_at=mapped.get("created_at") or now,
        )
        try:
            async with session.begin_nested():
                session.add(comment)
                await session.flush()
            imported += 1
        except IntegrityError:
            continue  # дедуп-гонка на partial unique index
    return imported


async def _ensure_talantix_consent(
    session: AsyncSession, company_id: UUID, candidate_id: UUID, pd: dict | None
) -> bool:
    """Создать Consent(status='signed') если в Talantix согласие 'agree' и его ещё нет."""
    if not talantix_mapper.pd_agreement_is_signed(pd):
        return False

    existing = (
        await session.execute(
            select(Consent.id).where(
                Consent.company_id == company_id,
                Consent.candidate_id == candidate_id,
            )
        )
    ).scalar_one_or_none()
    if existing:
        return False

    now = datetime.now(timezone.utc)
    signed_at = talantix_mapper.pd_agreement_signed_at(pd) or now
    number = await _generate_consent_number(session, company_id)
    consent = Consent(
        company_id=company_id,
        candidate_id=candidate_id,
        number=number,
        status="signed",
        channel="talantix",
        requested_at=signed_at,
        signed_at=signed_at,
        created_at=now,
    )
    try:
        async with session.begin_nested():
            session.add(consent)
            await session.flush()
        return True
    except IntegrityError:
        return False  # гонка на uq_consents_company_number


async def _process_talantix_person(
    session: AsyncSession,
    company_id: UUID,
    mapped: dict,
    events: list[dict],
    pd: dict | None,
    dedup_mode: str,
    stats: dict,
) -> None:
    """Обработать ОДНУ персону против переданной session (тестируемо с db_session).

    Создаёт/обогащает кандидата, импортирует комментарии (ВСЕГДА — главная ценность),
    создаёт Consent по согласию. Обновляет счётчики stats: created/updated/skipped/
    errors/comments. company_id ВЕЗДЕ."""
    first_name = (mapped.get("first_name") or "").strip()
    last_name = (mapped.get("last_name") or "").strip()
    phone = mapped.get("phone")
    email = mapped.get("email")
    tpid = _fit(mapped.get("talantix_person_id"), 64) if mapped.get("talantix_person_id") else None

    if not first_name and not last_name:
        stats["errors"] = stats.get("errors", 0) + 1
        return
    if not phone and not email:
        # Без контакта дедуп невозможен, кандидата не заводим — комментарии не к чему привязать.
        stats["errors"] = stats.get("errors", 0) + 1
        return

    existing = await _resolve_existing_talantix_candidate(session, company_id, tpid, phone, email)

    if existing is not None:
        candidate = existing
        if dedup_mode == "update":
            _apply_talantix_update(candidate, mapped)
            stats["updated"] = stats.get("updated", 0) + 1
        else:
            stats["skipped"] = stats.get("skipped", 0) + 1
        # Бэкфилл talantix_person_id для идемпотентности будущих импортов
        if tpid and not candidate.talantix_person_id:
            candidate.talantix_person_id = tpid
        await session.flush()
    else:
        candidate = Candidate(
            company_id=company_id,
            last_name=_fit(last_name, 120),
            first_name=_fit(first_name or "Неизвестно", 120),
            middle_name=_fit(mapped.get("middle_name"), 120) if mapped.get("middle_name") else None,
            phone=(normalize_phone(phone) or _fit(phone, 20)) if phone else None,
            email=_fit(email, 255) if email else None,
            city=_fit(mapped.get("city"), 120) if mapped.get("city") else None,
            birth_date=mapped.get("birth_date"),
            gender=mapped.get("gender"),
            last_position=_fit(mapped.get("last_position"), 255) if mapped.get("last_position") else None,
            resume_text=mapped.get("resume_text"),
            source="talantix",
            external_source="talantix",
            talantix_person_id=tpid,
            messengers=mapped.get("messengers") or [],
            extra={"imported": True},
        )
        session.add(candidate)
        await session.flush()
        stats["created"] = stats.get("created", 0) + 1

    # КОММЕНТАРИИ — импортируем для НОВОГО и СУЩЕСТВУЮЩЕГО кандидата (главная ценность)
    imported = await _import_talantix_comments(session, company_id, candidate.id, events)
    stats["comments"] = stats.get("comments", 0) + imported

    # Согласие 152-ФЗ
    await _ensure_talantix_consent(session, company_id, candidate.id, pd)

    # Кол-во откликов из Talantix — сохраняем в extra (без выдумывания структуры Response)
    rc = mapped.get("responses_count")
    if rc is not None:
        candidate.extra = {**(candidate.extra or {}), "talantix_responses_count": rc}
        await session.flush()


async def _run_talantix_import(job_id: UUID, company_id: UUID, user_id: UUID, dedup_mode: str):
    """Фоновое выполнение импорта из Talantix. company_id ВЕЗДЕ."""
    client = TalantixClient(company_id)
    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": 0, "comments": 0}
    try:
        logger.info("Начинаем импорт из Talantix company=%s", company_id)

        # Перечисляем все id персон (id-only, лёгкий запрос) → реальный total
        person_ids = await client.collect_person_ids(max_pages=TALANTIX_IMPORT_MAX_PAGES)
        await _update_job_progress(job_id, total=len(person_ids))

        processed = 0
        for pid in person_ids:
            try:
                node = await client.get_person_full(pid)
                if node is None:
                    # PersonError → пропускаем с логом, кандидат не создаётся
                    stats["errors"] += 1
                else:
                    mapped = talantix_mapper.map_person(node)
                    pd = node.get("personalDataAgreement")
                    events = await client.fetch_all_history(pid)
                    async with AsyncSessionLocal() as session:
                        await _process_talantix_person(
                            session, company_id, mapped, events, pd, dedup_mode, stats
                        )
                        await session.commit()
            except ValidationError:
                # Токен протух/интеграция отвалилась — тянуть нечего, прерываем весь импорт.
                raise
            except TalantixError as e:
                logger.warning("[talantix] персона %s: ошибка GraphQL, пропускаем: %s", pid, e)
                stats["errors"] += 1
            except Exception as e:  # noqa: BLE001
                logger.error("[talantix] персона %s: ошибка обработки: %s", pid, e)
                stats["errors"] += 1

            processed += 1
            if processed % 10 == 0 or processed == len(person_ids):
                await _update_job_progress(
                    job_id,
                    processed=processed,
                    created=stats["created"],
                    updated=stats["updated"],
                    skipped=stats["skipped"],
                    errors=stats["errors"],
                    comments_imported=stats["comments"],
                )

        # audit
        async with AsyncSessionLocal() as session:
            await audit(
                session,
                action="candidates_import_talantix",
                entity_type="candidate_import_job",
                entity_id=job_id,
                after={
                    "source": "talantix",
                    "created": stats["created"],
                    "updated": stats["updated"],
                    "skipped": stats["skipped"],
                    "errors": stats["errors"],
                    "comments_imported": stats["comments"],
                },
                actor_type="human",
                actor_user_id=user_id,
                company_id=company_id,
            )
            await session.commit()

        await _finalize_job(job_id, "done")

        try:
            await reindex_all_embeddings(company_id)
        except Exception as e:  # noqa: BLE001
            logger.error("Ошибка переиндексации после Talantix импорта: %s", e)

    except ValidationError as e:
        logger.error("Импорт Talantix %s прерван (токен/соединение): %s", job_id, e)
        await _finalize_job(job_id, "error", str(e)[:500])
    except Exception as e:  # noqa: BLE001
        logger.error("Ошибка выполнения импорта Talantix %s: %s", job_id, e)
        await _finalize_job(job_id, "error", str(e)[:500])
    finally:
        await client.close()


async def execute_talantix_import(
    session: AsyncSession, company_id: UUID, user_id: UUID, dedup_mode: str
) -> UUID:
    """Запуск импорта из Talantix в фоновой задаче (total уточняется в _run_talantix_import)."""
    job = await create_import_job(session, company_id, 0)
    await session.commit()

    task = asyncio.create_task(_run_talantix_import(job.id, company_id, user_id, dedup_mode))
    _active_tasks[job.id] = task

    def _cleanup_task(_t, _jid=job.id):
        _active_tasks.pop(_jid, None)

    task.add_done_callback(_cleanup_task)

    return job.id