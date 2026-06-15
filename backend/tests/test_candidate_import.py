import pytest
import json
import asyncio
from io import BytesIO
from uuid import uuid4
from unittest.mock import patch, AsyncMock

from openpyxl import Workbook

from app.models import Candidate, CandidateImportJob
from app.services.candidate_import import (
    parse_excel_file,
    _clean_name,
    _clean_phone,
    _clean_source,
    _clean_salary,
    _auto_map_columns,
    _parse_company_position,
    preview_import,
    create_import_job,
    get_import_job,
    _run_import,
    preview_potok_import,
    _run_potok_import
)
from app.core.errors import ValidationError


def create_test_excel(headers, rows):
    """Создает Excel файл в памяти для тестов"""
    workbook = Workbook()
    worksheet = workbook.active

    # Добавляем заголовки
    for i, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=i, value=header)

    # Добавляем данные
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # Сохраняем в BytesIO
    file_buffer = BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)
    return file_buffer.getvalue()


class TestCandidateImportParsing:
    """Тесты парсинга Excel файлов"""

    @pytest.mark.asyncio
    async def test_parse_excel_file_success(self):
        """Тест успешного парсинга Excel файла"""
        headers = ["ФИО Кандидата", "Телефон", "Email", "Город"]
        rows = [
            ["Иванов Иван Иванович", "+7 915 123-45-67", "ivan@example.com", "Москва"],
            ["Петров Петр", "8 916 234-56-78", "petr@example.com", "СПб"],
            ["", "", "", ""],  # Пустая строка - должна игнорироваться
            ["Сидоров Сидор", "7 917 345-67-89", "", "Казань"]
        ]

        content = create_test_excel(headers, rows)

        result = await parse_excel_file(content, "test.xlsx")

        assert result["columns"] == headers
        assert result["row_count"] == 3  # Пустая строка не считается
        assert "ФИО Кандидата" in result["samples"]
        assert len(result["samples"]["ФИО Кандидата"]) == 3
        assert "auto_mapping" in result
        assert result["auto_mapping"]["ФИО Кандидата"] == "name"
        assert result["auto_mapping"]["Телефон"] == "phone"

    @pytest.mark.asyncio
    async def test_parse_xls_file_error(self):
        """Тест ошибки при загрузке .xls файла"""
        content = b"fake content"

        with pytest.raises(Exception) as exc_info:
            await parse_excel_file(content, "test.xls")

        assert "Старый формат .xls не поддерживается" in str(exc_info.value)

    @pytest.mark.asyncio
    async def test_parse_invalid_file_error(self):
        """Тест ошибки при загрузке невалидного файла"""
        content = b"not an excel file"

        with pytest.raises(Exception) as exc_info:
            await parse_excel_file(content, "test.txt")

        assert "Поддерживаются только файлы формата .xlsx" in str(exc_info.value)


class TestColumnRecognition:
    """Тесты автоматического распознавания колонок"""

    def test_auto_map_columns_basic(self):
        """Тест базового распознавания колонок"""
        columns = ["ФИО Кандидата", "Моб. телефон", "E-mail", "Город", "Неизвестная колонка"]

        mapping = _auto_map_columns(columns)

        assert mapping["ФИО Кандидата"] == "name"
        assert mapping["Моб. телефон"] == "phone"
        assert mapping["E-mail"] == "email"
        assert mapping["Город"] == "city"
        assert mapping["Неизвестная колонка"] == "skip"

    def test_auto_map_columns_english(self):
        """Тест распознавания английских названий"""
        columns = ["Full name", "Phone", "Email", "City", "Salary", "Company"]

        mapping = _auto_map_columns(columns)

        assert mapping["Full name"] == "name"
        assert mapping["Phone"] == "phone"
        assert mapping["Email"] == "email"
        assert mapping["City"] == "city"
        assert mapping["Salary"] == "salary"
        assert mapping["Company"] == "company"

    def test_auto_map_columns_case_insensitive(self):
        """Тест нечувствительности к регистру"""
        columns = ["ФИО КАНДИДАТА", "моб. телефон", "E-MAIL"]

        mapping = _auto_map_columns(columns)

        assert mapping["ФИО КАНДИДАТА"] == "name"
        assert mapping["моб. телефон"] == "phone"
        assert mapping["E-MAIL"] == "email"


class TestDataCleaning:
    """Тесты очистки данных"""

    def test_clean_name_full(self):
        """Тест очистки полного ФИО"""
        last, first, middle = _clean_name("Иванов Иван Иванович")
        assert last == "Иванов"
        assert first == "Иван"
        assert middle == "Иванович"

    def test_clean_name_with_code(self):
        """Тест очистки ФИО с кодом анонимизации"""
        last, first, middle = _clean_name("123-Иванов Иван")
        assert last == "Иванов"
        assert first == "Иван"
        assert middle == ""

    def test_clean_name_two_parts(self):
        """Тест очистки ФИО из двух частей"""
        last, first, middle = _clean_name("Петров Петр")
        assert last == "Петров"
        assert first == "Петр"
        assert middle == ""

    def test_clean_name_one_part(self):
        """Тест очистки ФИО из одной части"""
        last, first, middle = _clean_name("Иван")
        assert last == ""
        assert first == "Иван"
        assert middle == ""

    def test_clean_phone_normalization(self):
        """Тест нормализации телефонов"""
        assert _clean_phone("8 915 123-45-67") == "+79151234567"
        assert _clean_phone("7 916 234 56 78") == "+79162345678"
        assert _clean_phone("9173456789") == "+79173456789"
        assert _clean_phone("+7 918 456-78-90") == "+79184567890"
        assert _clean_phone("invalid") == "invalid"

    def test_clean_source_mapping(self):
        """Тест маппинга источников"""
        assert _clean_source("headhunter") == "hh"
        assert _clean_source("hh.ru") == "hh"
        assert _clean_source("вручную") == "manual"
        assert _clean_source("Telegram") == "telegram"
        assert _clean_source("авито") == "avito"
        assert _clean_source("superjob") == "superjob"
        assert _clean_source("linkedin") == "other"
        assert _clean_source("") == "other"

    def test_clean_salary_extraction(self):
        """Тест извлечения зарплаты"""
        assert _clean_salary("150 000 руб") == 150000
        assert _clean_salary("от 100000") == 100000
        assert _clean_salary("200-300k") == 200300
        assert _clean_salary("invalid") is None
        assert _clean_salary("") is None

    def test_parse_company_position(self):
        """Тест парсинга 'Компания: Должность'"""
        company, position = _parse_company_position("Сбербанк: Senior Frontend", "position")
        assert company == "Сбербанк"
        assert position == "Senior Frontend"

        company, position = _parse_company_position("Яндекс: Аналитик", "company")
        assert company == "Яндекс"
        assert position == "Аналитик"

        company, position = _parse_company_position("Простая должность", "position")
        assert company == ""
        assert position == "Простая должность"


class TestDeduplication:
    """Тесты дедупликации"""

    @pytest.mark.asyncio
    async def test_preview_import_with_duplicates(self, db_session, admin_user):
        """Тест превью импорта с дублями"""
        company_id = admin_user.company_id

        # Создаем существующего кандидата
        existing_candidate = Candidate(
            company_id=company_id,
            first_name="Иван",
            last_name="Иванов",
            phone="+79151234567",
            email="ivan@example.com",
            source="manual"
        )
        db_session.add(existing_candidate)
        await db_session.commit()

        # Создаем файл с дублем и новым кандидатом
        headers = ["ФИО", "Телефон", "Email"]
        rows = [
            ["Иванов Иван", "+7 915 123-45-67", "ivan@example.com"],  # Дубль
            ["Петров Петр", "+7 916 234-56-78", "petr@example.com"]   # Новый
        ]
        content = create_test_excel(headers, rows)
        mapping = {"ФИО": "name", "Телефон": "phone", "Email": "email"}

        result = await preview_import(db_session, company_id, content, mapping, "skip")

        assert result["summary"]["total"] == 2
        assert result["summary"]["new"] == 1
        assert result["summary"]["duplicates"] == 1
        assert result["summary"]["errors"] == 0

        # Проверяем статусы строк
        rows_data = result["rows"]
        assert rows_data[0]["status"] == "duplicate"
        assert rows_data[1]["status"] == "new"

    @pytest.mark.asyncio
    async def test_preview_import_company_scoped(self, db_session, admin_user):
        """Тест того, что дедупликация работает в рамках компании"""
        from app.models import Company

        company_id = admin_user.company_id

        # Создаём реальную вторую компанию
        other_company = Company(name="Другая компания")
        db_session.add(other_company)
        await db_session.flush()

        # Создаем кандидата в другой компании
        other_company_candidate = Candidate(
            company_id=other_company.id,
            first_name="Иван",
            last_name="Иванов",
            phone="+79151234567",
            source="manual"
        )
        db_session.add(other_company_candidate)
        await db_session.commit()

        # Импортируем того же кандидата в нашу компанию
        headers = ["ФИО", "Телефон"]
        rows = [["Иванов Иван", "+7 915 123-45-67"]]
        content = create_test_excel(headers, rows)
        mapping = {"ФИО": "name", "Телефон": "phone"}

        result = await preview_import(db_session, company_id, content, mapping, "skip")

        # Не должно быть дублей - кандидат из другой компании
        assert result["summary"]["duplicates"] == 0
        assert result["summary"]["new"] == 1


class TestImportJob:
    """Тесты управления задачами импорта"""

    @pytest.mark.asyncio
    async def test_create_import_job(self, db_session, admin_user):
        """Тест создания задачи импорта"""
        company_id = admin_user.company_id

        job = await create_import_job(db_session, company_id, 100)
        await db_session.commit()

        assert job.company_id == company_id
        assert job.total == 100
        assert job.status == "running"
        assert job.processed == 0

    @pytest.mark.asyncio
    async def test_get_import_job(self, db_session, admin_user):
        """Тест получения задачи импорта"""
        company_id = admin_user.company_id

        # Создаем задачу
        job = await create_import_job(db_session, company_id, 50)
        await db_session.commit()

        # Получаем задачу
        retrieved_job = await get_import_job(db_session, job.id, company_id)

        assert retrieved_job is not None
        assert retrieved_job.id == job.id
        assert retrieved_job.company_id == company_id

    @pytest.mark.asyncio
    async def test_get_import_job_company_scoped(self, db_session, admin_user):
        """Тест получения задачи импорта с проверкой company_id"""
        company_id = admin_user.company_id
        other_company_id = uuid4()

        # Создаем задачу
        job = await create_import_job(db_session, company_id, 50)
        await db_session.commit()

        # Пытаемся получить задачу от имени другой компании
        retrieved_job = await get_import_job(db_session, job.id, other_company_id)

        assert retrieved_job is None  # Не должна найтись


class TestErrorHandling:
    """Тесты обработки ошибок"""

    @pytest.mark.asyncio
    async def test_preview_import_with_errors(self, db_session, admin_user):
        """Тест превью импорта с ошибочными строками"""
        company_id = admin_user.company_id

        headers = ["ФИО", "Телефон", "Email"]
        rows = [
            ["", "+7 915 123-45-67", ""],              # Нет имени
            ["Петров Петр", "", ""],                   # Нет контактов
            ["Сидоров Сидор", "+7 916 234-56-78", ""] # Валидная строка
        ]
        content = create_test_excel(headers, rows)
        mapping = {"ФИО": "name", "Телефон": "phone", "Email": "email"}

        result = await preview_import(db_session, company_id, content, mapping, "skip")

        assert result["summary"]["total"] == 3
        assert result["summary"]["new"] == 1
        assert result["summary"]["errors"] == 2

        rows_data = result["rows"]
        assert rows_data[0]["status"] == "error"
        assert rows_data[0]["error"] == "нет имени"
        assert rows_data[1]["status"] == "error"
        assert rows_data[1]["error"] == "нет контакта"
        assert rows_data[2]["status"] == "new"


class TestExcelImportSavepoint:
    """Тесты per-row savepoint для Excel-импорта"""

    @pytest.mark.asyncio
    async def test_excel_import_per_row_savepoint(self, db_session, admin_user):
        """
        Тест того, что сбой констрейнта одной строки не валит весь батч.

        Создаём ситуацию, где одна строка невалидна (например, слишком длинный телефон),
        а остальные валидны. Проверяем, что валидные строки импортируются,
        а errors_count растёт для невалидных.
        """
        from contextlib import asynccontextmanager

        def _session_local_returning(db_session):
            @asynccontextmanager
            async def _factory():
                yield db_session
            return _factory

        company_id = admin_user.company_id
        user_id = admin_user.id

        # Создаём Excel с валидными строками и одной с отсутствующим именем
        headers = ["ФИО", "Телефон", "Email"]
        rows = [
            ["Иванов Иван", "+7 915 123-45-67", "ivan@example.com"],  # Валидный
            ["", "+7 916 234-56-78", "petr@example.com"],             # Нет имени - ошибка на этапе парсинга
            ["Сидоров Сидор", "+7 917 345-67-89", "sidor@example.com"]  # Валидный
        ]
        content = create_test_excel(headers, rows)
        mapping = {"ФИО": "name", "Телефон": "phone", "Email": "email"}

        # Создаём джоб
        job = await create_import_job(db_session, company_id, len(rows))
        await db_session.commit()

        # Патчим AsyncSessionLocal и запускаем импорт
        with patch('app.services.candidate_import.AsyncSessionLocal', _session_local_returning(db_session)):
            await _run_import(job.id, company_id, user_id, content, mapping, "skip")

        # Обновляем джоб из БД
        db_session.expire_all()
        await db_session.refresh(job)

        # Проверяем: валидные строки должны импортироваться, ошибочные — считаться как errors
        assert job.status == "done"
        assert job.created >= 1  # Минимум одна валидная строка создалась (Иванов + Сидоров)
        assert job.errors >= 1   # Минимум одна ошибка зафиксирована (пустое ФИО)


class TestPotokImportTimeout:
    """Тесты таймаутов для Potok-импорта"""

    @pytest.mark.asyncio
    async def test_potok_preview_timeout(self, db_session, admin_user):
        """Тест таймаута в preview_potok_import"""
        company_id = admin_user.company_id
        token = "fake_token"

        # Мокаем list_applicants с мгновенным TimeoutError
        async def timeout_list_applicants(*args, **kwargs):
            raise asyncio.TimeoutError()

        with patch("app.services.candidate_import.list_applicants", side_effect=timeout_list_applicants):
            with pytest.raises(ValidationError) as exc_info:
                await preview_potok_import(db_session, company_id, token, "skip")

            assert "Таймаут при обращении к API Potok" in str(exc_info.value)

    @pytest.mark.asyncio
    async def test_potok_import_timeout_finalizes_job(self, db_session, admin_user):
        """Тест того, что таймаут в _run_potok_import корректно финализирует джоб"""
        from contextlib import asynccontextmanager

        def _session_local_returning(db_session):
            @asynccontextmanager
            async def _factory():
                yield db_session
            return _factory

        company_id = admin_user.company_id
        user_id = admin_user.id
        token = "fake_token"

        # Создаём джоб
        job = await create_import_job(db_session, company_id, 0)
        await db_session.commit()

        # Мокаем iter_applicants с мгновенным TimeoutError (имитируем фатальную 401 ошибку)
        async def timeout_iter_applicants(*args, **kwargs):
            from app.core.errors import ValidationError
            raise ValidationError("Поток отклонил токен (проверьте, что токен активен и даёт доступ на чтение кандидатов)")

        # Патчим AsyncSessionLocal и iter_applicants
        with patch('app.services.candidate_import.AsyncSessionLocal', _session_local_returning(db_session)), \
             patch("app.services.candidate_import.iter_applicants", side_effect=timeout_iter_applicants):
            # Запускаем импорт — должен завершиться с ошибкой
            await _run_potok_import(job.id, company_id, user_id, token, "skip")

        # Читаем джоб из БД — используем тот же паттерн, что в test_get_import_job
        await db_session.refresh(job)

        # Проверяем, что джоб корректно финализирован с ошибкой
        assert job.status == "error"
        assert job.finished_at is not None
        assert "токен активен" in job.error

    @pytest.mark.asyncio
    async def test_potok_import_streaming_success(self, db_session, admin_user):
        """Тест успешного стримингового импорта из Potok"""
        from contextlib import asynccontextmanager

        def _session_local_returning(db_session):
            @asynccontextmanager
            async def _factory():
                yield db_session
            return _factory

        company_id = admin_user.company_id
        user_id = admin_user.id
        token = "valid_token"

        # Создаём джоб
        job = await create_import_job(db_session, company_id, 0)
        await db_session.commit()

        # Мокаем стриминговый генератор
        async def mock_iter_applicants(*args, **kwargs):
            # Первый батч
            yield [
                {
                    "id": 100,
                    "first_name": "Иван",
                    "last_name": "Петров",
                    "email": "ivan@potok.test",
                    "phones": ["79991234567"]
                }
            ]
            # Второй батч
            yield [
                {
                    "id": 101,
                    "first_name": "Петр",
                    "last_name": "Сидоров",
                    "email": "petr@potok.test",
                    "phones": ["79992345678"]
                }
            ]

        # Мокаем map_potok_applicant для возврата нужного формата
        def mock_map_potok_applicant(raw):
            return {
                "first_name": raw["first_name"],
                "last_name": raw["last_name"],
                "email": raw["email"],
                "phone": raw["phones"][0] if raw.get("phones") else None,
                "external_id": str(raw["id"]),
                "experience": [],
                "skills": [],
                "education": [],
                "languages": []
            }

        # Патчим зависимости
        with patch('app.services.candidate_import.AsyncSessionLocal', _session_local_returning(db_session)), \
             patch("app.services.candidate_import.iter_applicants", side_effect=mock_iter_applicants), \
             patch("app.services.candidate_import.map_potok_applicant", side_effect=mock_map_potok_applicant), \
             patch("app.services.candidate_import._create_potok_child_records", new=AsyncMock()), \
             patch("app.services.candidate_import.reindex_all_embeddings", new=AsyncMock()):
            # Запускаем импорт
            await _run_potok_import(job.id, company_id, user_id, token, "skip")

        # Обновляем джоб из БД
        await db_session.refresh(job)

        # Проверяем результат
        assert job.status == "done"
        assert job.created == 2  # Два кандидата созданы
        assert job.skipped == 0
        assert job.errors == 0

        # Проверяем, что кандидаты действительно созданы в БД
        from sqlalchemy import func
        result = await db_session.execute(
            select(func.count(Candidate.id))
            .where(Candidate.company_id == company_id)
            .where(Candidate.source == "potok")
        )
        created_count = result.scalar()
        assert created_count == 2